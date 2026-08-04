import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from api.agent_api import run_agent



REPORT_DIR = "reports"


os.makedirs(
    REPORT_DIR,
    exist_ok=True
)



def create_excel_report(question):


    # ======================
    # 调用Agent
    # ======================

    result = run_agent(
        question
    )


    if result["status"] != "success":

        raise Exception(
            result["message"]
        )


    dashboard = result["dashboard"]



    file_path = os.path.join(

        REPORT_DIR,

        "sales_report.xlsx"

    )



    wb = Workbook()



    # ======================
    # Sheet1 KPI
    # ======================


    ws = wb.active

    ws.title = "经营指标"



    ws.append(
        [
            "指标",
            "数值"
        ]
    )


    kpi = dashboard.get(
        "kpi",
        {}
    )


    rows = [

        [
            "总销售额",
            kpi.get(
                "total_sales",
                0
            )
        ],

        [
            "总利润",
            kpi.get(
                "total_profit",
                0
            )
        ],

        [
            "平均月销售",
            kpi.get(
                "average_monthly_sales",
                0
            )
        ],

        [
            "平均利润率",
            kpi.get(
                "average_profit_margin",
                0
            )
        ]

    ]


    for row in rows:

        ws.append(row)




    # ======================
    # Sheet2 Top10客户
    # ======================


    ws2 = wb.create_sheet(
        "Top10客户"
    )


    ws2.append(

        [
            "客户",
            "消费金额",
            "购买次数",
            "RFM",
            "CLV"
        ]

    )



    customers = (

        dashboard
        .get(
            "customer",
            {}
        )
        .get(
            "top10_customer",
            []
        )

    )



    for c in customers:


        ws2.append(

            [

                c.get(
                    "Customer Name"
                ),

                c.get(
                    "Monetary"
                ),

                c.get(
                    "Frequency"
                ),

                c.get(
                    "RFM_Score"
                ),

                c.get(
                    "CLV"
                )

            ]

        )




    # ======================
    # Sheet3 客户分层
    # ======================


    ws3 = wb.create_sheet(
        "客户分层"
    )


    ws3.append(
        [
            "客户类型",
            "数量"
        ]
    )


    level = (

        dashboard
        .get(
            "customer",
            {}
        )
        .get(
            "level_distribution",
            {}
        )

    )


    for k,v in level.items():

        ws3.append(
            [
                k,
                v
            ]
        )



    # ======================
    # 自动调整
    # ======================


    for ws in wb:


        for col in ws.columns:

            max_length = max(

                len(
                    str(cell.value)
                )

                if cell.value

                else 0

                for cell in col

            )


            ws.column_dimensions[

                col[0].column_letter

            ].width = max_length + 5



        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center"
            )



    wb.save(
        file_path
    )


    return file_path